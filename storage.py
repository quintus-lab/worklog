"""SQLite primary store + Excel export/backup/restore (stdlib sqlite3 + openpyxl).

No admin rights required: database is a single file under data/.
"""

from __future__ import annotations

import shutil
import sqlite3
import threading
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import tickets

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DB_PATH = DATA_DIR / "worklog.db"
EXCEL_PATH = DATA_DIR / "work_log.xlsx"
BACKUP_DIR = DATA_DIR / "backups"
LEGACY_XLSX = DATA_DIR / "work_log.xlsx"

MAX_BACKUPS = 40
MAX_TITLE_LEN = 200
MAX_DETAILS_LEN = 8000
MAX_TAGS_LEN = 200
MAX_SEARCH_Q_LEN = 200

CATEGORIES = [
    "General",
    "Network",
    "Incident",
    "Project",
    "Meeting",
    "Change",
    "Documentation",
    "Learning",
    "Other",
]
STATUSES = ["done", "in-progress", "blocked", "follow-up"]

HEADERS = [
    "ID",
    "Date",
    "Title",
    "Details",
    "Category",
    "Status",
    "Tags",
    "Follow Up",
    "Owner",
    "Created At",
    "Updated At",
]

_lock = threading.RLock()
_initialized = False


class StorageError(Exception):
    """Raised when save/export fails (e.g. Excel file locked)."""


def clip(s: str, n: int) -> str:
    s = s or ""
    return s if len(s) <= n else s[:n]


def _require_ymd(value: str, label: str) -> str:
    raw = (value or "").strip()[:10]
    try:
        datetime.strptime(raw, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"{label} must be YYYY-MM-DD") from exc
    return raw


def normalize_entry_fields(data: dict[str, Any], *, partial: bool = False) -> dict[str, Any]:
    """Validate and normalize entry fields. Raises ValueError on bad input."""
    out: dict[str, Any] = {}
    if not partial or "title" in data:
        title = clip(str(data.get("title") or "").strip(), MAX_TITLE_LEN)
        if not title:
            raise ValueError("Title is required")
        out["title"] = title
    if not partial or "details" in data:
        out["details"] = clip(str(data.get("details") or "").strip(), MAX_DETAILS_LEN)
    if not partial or "date" in data:
        out["date"] = _require_ymd(
            str(data.get("date") or date.today().isoformat()), "Date"
        )
    if not partial or "category" in data:
        cat = str(data.get("category") or "General").strip() or "General"
        out["category"] = cat if cat in CATEGORIES else "Other"
    if not partial or "status" in data:
        st = str(data.get("status") or "done").strip() or "done"
        out["status"] = st if st in STATUSES else "done"
    if not partial or "tags" in data:
        out["tags"] = clip(str(data.get("tags") or "").strip(), MAX_TAGS_LEN)
    if not partial or "follow_up" in data:
        fu = str(data.get("follow_up") or "").strip()[:10]
        out["follow_up"] = _require_ymd(fu, "Follow-up") if fu else ""
    if not partial or "owner" in data:
        out["owner"] = str(data.get("owner") or "").strip()
    return out


def _connect() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH), timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db() -> None:
    """Create schema and migrate legacy Excel once."""
    global _initialized
    with _lock:
        if _initialized and DB_PATH.exists():
            return
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        conn = _connect()
        try:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS entries (
                    id TEXT PRIMARY KEY,
                    date TEXT NOT NULL,
                    title TEXT NOT NULL,
                    details TEXT NOT NULL DEFAULT '',
                    category TEXT NOT NULL DEFAULT 'General',
                    status TEXT NOT NULL DEFAULT 'done',
                    tags TEXT NOT NULL DEFAULT '',
                    follow_up TEXT NOT NULL DEFAULT '',
                    owner TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_entries_date ON entries(date);
                CREATE INDEX IF NOT EXISTS idx_entries_status ON entries(status);
                CREATE INDEX IF NOT EXISTS idx_entries_owner ON entries(owner);

                CREATE TABLE IF NOT EXISTS meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                );
                """
            )
            conn.commit()
            # One-time import from Excel if DB empty and xlsx has rows
            n = conn.execute("SELECT COUNT(*) AS c FROM entries").fetchone()["c"]
            if n == 0 and LEGACY_XLSX.is_file():
                imported = _import_xlsx_rows(conn, LEGACY_XLSX)
                if imported:
                    _set_meta(conn, "migrated_from_xlsx", datetime.now().isoformat(timespec="seconds"))
                    conn.commit()
                    # Keep a pre-migration backup of the xlsx
                    try:
                        backup_excel_copy(LEGACY_XLSX, label="pre-sqlite-migrate")
                    except Exception:
                        pass
        finally:
            conn.close()
        _initialized = True


def _set_meta(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        "INSERT INTO meta(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )


def get_meta(key: str, default: str = "") -> str:
    init_db()
    with _lock:
        conn = _connect()
        try:
            row = conn.execute(
                "SELECT value FROM meta WHERE key = ?", (key,)
            ).fetchone()
            return str(row["value"]) if row else default
        finally:
            conn.close()


def get_meta_many(keys: list[str]) -> dict[str, str]:
    """Read several meta keys in one connection."""
    init_db()
    if not keys:
        return {}
    with _lock:
        conn = _connect()
        try:
            out = {k: "" for k in keys}
            placeholders = ",".join("?" for _ in keys)
            rows = conn.execute(
                f"SELECT key, value FROM meta WHERE key IN ({placeholders})",
                keys,
            ).fetchall()
            for row in rows:
                out[str(row["key"])] = str(row["value"] or "")
            return out
        finally:
            conn.close()


def set_meta(key: str, value: str) -> None:
    init_db()
    with _lock:
        conn = _connect()
        try:
            _set_meta(conn, key, value)
            conn.commit()
        finally:
            conn.close()


def get_ticket_settings() -> dict[str, str]:
    meta = get_meta_many(["ticket_url", "ticket_prefixes"])
    return {
        "url": meta.get("ticket_url", ""),
        "prefixes": meta.get("ticket_prefixes", ""),
        "default_prefixes": ",".join(tickets.DEFAULT_PREFIXES),
    }


def save_ticket_settings(url: str, prefixes: str) -> dict[str, str]:
    url, prefixes = tickets.validate_ticket_settings(url, prefixes)
    init_db()
    with _lock:
        conn = _connect()
        try:
            _set_meta(conn, "ticket_url", url)
            _set_meta(conn, "ticket_prefixes", prefixes)
            conn.commit()
        finally:
            conn.close()
    return get_ticket_settings()


def clip_search_q(q: str | None) -> str | None:
    """Normalize a search string; None/blank → None. Caps length."""
    if q is None:
        return None
    raw = str(q).strip()
    if not raw:
        return None
    return raw[:MAX_SEARCH_Q_LEN]


def _like_contains(q: str) -> str:
    """LIKE pattern for substring match; % _ \\ escaped."""
    escaped = (
        q.replace("\\", "\\\\")
        .replace("%", "\\%")
        .replace("_", "\\_")
    )
    return f"%{escaped}%"


def _search_clause(q: str) -> tuple[str, list[Any]]:
    like = _like_contains(q)
    sql = (
        " AND (title LIKE ? ESCAPE '\\' OR details LIKE ? ESCAPE '\\' "
        "OR tags LIKE ? ESCAPE '\\' OR category LIKE ? ESCAPE '\\' "
        "OR owner LIKE ? ESCAPE '\\' OR status LIKE ? ESCAPE '\\')"
    )
    return sql, [like, like, like, like, like, like]


def _row_to_entry(row: sqlite3.Row | dict) -> dict[str, Any]:
    if isinstance(row, sqlite3.Row):
        d = dict(row)
    else:
        d = dict(row)
    return {
        "id": str(d.get("id") or ""),
        "date": str(d.get("date") or "")[:10],
        "title": str(d.get("title") or ""),
        "details": str(d.get("details") or ""),
        "category": str(d.get("category") or "General"),
        "status": str(d.get("status") or "done"),
        "tags": str(d.get("tags") or ""),
        "follow_up": str(d.get("follow_up") or "")[:10],
        "owner": str(d.get("owner") or ""),
        "created_at": str(d.get("created_at") or ""),
        "updated_at": str(d.get("updated_at") or ""),
    }


def _parse_xlsx_entries(path: Path, *, legacy_9col: bool = False) -> list[dict[str, Any]]:
    """Parse work log rows from Excel. Prefers header row; optional legacy 9-col layout."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    has_header = bool(header and header[0].upper() == "ID")
    out: list[dict[str, Any]] = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if has_header:
        idx = {h.lower(): i for i, h in enumerate(header)}

        def g(vals: list, *keys: str, default: str = "") -> str:
            for k in keys:
                i = idx.get(k.lower())
                if i is not None and i < len(vals) and vals[i] is not None:
                    return str(vals[i])
            return default

        for r in rows[1:]:
            if not r or r[0] in (None, "ID"):
                continue
            vals = list(r)
            entry = {
                "id": g(vals, "id") or str(uuid.uuid4()),
                "date": g(vals, "date")[:10],
                "title": g(vals, "title"),
                "details": g(vals, "details"),
                "category": g(vals, "category", default="General") or "General",
                "status": g(vals, "status", default="done") or "done",
                "tags": g(vals, "tags", "tag", "ticket"),
                "follow_up": g(vals, "follow up", "follow_up", "follow-up")[:10],
                "owner": g(vals, "owner"),
                "created_at": g(vals, "created at", "created_at") or now,
                "updated_at": g(vals, "updated at", "updated_at") or now,
            }
            if entry["title"]:
                out.append(entry)
        return out

    # No header: explicit layouts only
    for r in rows[1 if has_header else 0 :]:
        if not r or r[0] in (None, "ID"):
            continue
        vals = list(r)
        while len(vals) < 11:
            vals.append(None)
        if legacy_9col or (len([v for v in vals if v is not None]) <= 9 and vals[6] is not None):
            # ID Date Title Details Category Status Owner Created Updated
            entry = {
                "id": str(vals[0] or uuid.uuid4()),
                "date": str(vals[1] or "")[:10],
                "title": str(vals[2] or ""),
                "details": str(vals[3] or ""),
                "category": str(vals[4] or "General"),
                "status": str(vals[5] or "done"),
                "tags": "",
                "follow_up": "",
                "owner": str(vals[6] or ""),
                "created_at": str(vals[7] or now),
                "updated_at": str(vals[8] or now),
            }
        else:
            # ID Date Title Details Category Status Tags FollowUp Owner Created Updated
            entry = {
                "id": str(vals[0] or uuid.uuid4()),
                "date": str(vals[1] or "")[:10],
                "title": str(vals[2] or ""),
                "details": str(vals[3] or ""),
                "category": str(vals[4] or "General"),
                "status": str(vals[5] or "done"),
                "tags": str(vals[6] or ""),
                "follow_up": str(vals[7] or "")[:10],
                "owner": str(vals[8] or ""),
                "created_at": str(vals[9] or now),
                "updated_at": str(vals[10] or now),
            }
        if entry["title"]:
            out.append(entry)
    return out


def _upsert_entry(conn: sqlite3.Connection, entry: dict[str, Any], *, mode: str) -> str:
    """mode: insert | upsert. Returns 'inserted'|'updated'|'skipped'."""
    exists = conn.execute("SELECT 1 FROM entries WHERE id=?", (entry["id"],)).fetchone()
    if exists and mode == "insert":
        return "skipped"
    if exists:
        conn.execute(
            """
            UPDATE entries SET date=?, title=?, details=?, category=?, status=?,
              tags=?, follow_up=?, owner=?, created_at=?, updated_at=?
            WHERE id=?
            """,
            (
                entry["date"],
                entry["title"],
                entry["details"],
                entry["category"],
                entry["status"],
                entry["tags"],
                entry["follow_up"],
                entry["owner"],
                entry["created_at"],
                entry["updated_at"],
                entry["id"],
            ),
        )
        return "updated"
    conn.execute(
        """
        INSERT INTO entries
        (id, date, title, details, category, status, tags, follow_up, owner, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            entry["id"],
            entry["date"],
            entry["title"],
            entry["details"],
            entry["category"],
            entry["status"],
            entry["tags"],
            entry["follow_up"],
            entry["owner"],
            entry["created_at"],
            entry["updated_at"],
        ),
    )
    return "inserted"


def import_excel(
    path: Path | str,
    *,
    mode: str = "insert",
    legacy_9col: bool = False,
    copy_to_data: bool = True,
) -> dict[str, Any]:
    """Import entries from .xlsx into SQLite.

    mode:
      insert  — skip existing IDs
      upsert  — update existing IDs
      replace — backup current DB, wipe entries, then insert all from Excel
    """
    if mode not in ("insert", "upsert", "replace"):
        raise ValueError("mode must be insert, upsert, or replace")
    src = Path(path).expanduser().resolve()
    if not src.is_file():
        raise FileNotFoundError(f"Excel file not found: {src}")

    init_db()
    before = count_entries()
    if copy_to_data and src.resolve() != EXCEL_PATH.resolve():
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, EXCEL_PATH)

    entries = _parse_xlsx_entries(src, legacy_9col=legacy_9col)
    if mode == "replace":
        create_backup(reason="pre-excel-import-replace")
        with _lock:
            conn = _connect()
            try:
                conn.execute("DELETE FROM entries")
                conn.commit()
            finally:
                conn.close()

    inserted = updated = skipped = 0
    write_mode = "insert" if mode == "replace" else mode
    with _lock:
        conn = _connect()
        try:
            for entry in entries:
                result = _upsert_entry(conn, entry, mode=write_mode)
                if result == "inserted":
                    inserted += 1
                elif result == "updated":
                    updated += 1
                else:
                    skipped += 1
            conn.commit()
        finally:
            conn.close()

    try:
        export_excel(EXCEL_PATH)
    except StorageError:
        pass

    return {
        "before": before,
        "after": count_entries(),
        "parsed": len(entries),
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "mode": mode,
    }


def _import_xlsx_rows(conn: sqlite3.Connection, path: Path) -> int:
    """Legacy helper for first-boot migrate; inserts new IDs only."""
    n = 0
    for entry in _parse_xlsx_entries(path):
        if _upsert_entry(conn, entry, mode="insert") == "inserted":
            n += 1
    return n


def load_entries(
    *,
    start: str | None = None,
    end: str | None = None,
    q: str | None = None,
    status: str | None = None,
    limit: int | None = None,
    offset: int = 0,
) -> list[dict]:
    init_db()
    q = clip_search_q(q)
    sql = "SELECT * FROM entries WHERE 1=1"
    params: list[Any] = []
    if start:
        sql += " AND date >= ?"
        params.append(start)
    if end:
        sql += " AND date <= ?"
        params.append(end)
    if status:
        sql += " AND status = ?"
        params.append(status)
    if q:
        clause, like_params = _search_clause(q)
        sql += clause
        params.extend(like_params)
    sql += " ORDER BY date DESC, created_at DESC"
    if limit is not None:
        sql += " LIMIT ? OFFSET ?"
        params.extend([int(limit), int(offset)])
    with _lock:
        conn = _connect()
        try:
            rows = conn.execute(sql, params).fetchall()
            return [_row_to_entry(r) for r in rows]
        finally:
            conn.close()


def count_entries(
    *,
    start: str | None = None,
    end: str | None = None,
    q: str | None = None,
    status: str | None = None,
) -> int:
    init_db()
    q = clip_search_q(q)
    sql = "SELECT COUNT(*) AS c FROM entries WHERE 1=1"
    params: list[Any] = []
    if start:
        sql += " AND date >= ?"
        params.append(start)
    if end:
        sql += " AND date <= ?"
        params.append(end)
    if status:
        sql += " AND status = ?"
        params.append(status)
    if q:
        clause, like_params = _search_clause(q)
        sql += clause
        params.extend(like_params)
    with _lock:
        conn = _connect()
        try:
            return int(conn.execute(sql, params).fetchone()["c"])
        finally:
            conn.close()


def get_entry(entry_id: str) -> dict | None:
    init_db()
    with _lock:
        conn = _connect()
        try:
            row = conn.execute(
                "SELECT * FROM entries WHERE id = ?", (entry_id,)
            ).fetchone()
            return _row_to_entry(row) if row else None
        finally:
            conn.close()


def create_entry(data: dict[str, Any]) -> dict:
    init_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fields = normalize_entry_fields(data, partial=False)
    entry = {
        "id": str(data.get("id") or uuid.uuid4()),
        **fields,
        "created_at": now,
        "updated_at": now,
    }
    with _lock:
        conn = _connect()
        try:
            conn.execute(
                """
                INSERT INTO entries
                (id, date, title, details, category, status, tags, follow_up, owner, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    entry["id"],
                    entry["date"],
                    entry["title"],
                    entry["details"],
                    entry["category"],
                    entry["status"],
                    entry["tags"],
                    entry["follow_up"],
                    entry["owner"],
                    entry["created_at"],
                    entry["updated_at"],
                ),
            )
            conn.commit()
        finally:
            conn.close()
    _after_write()
    return entry


def update_entry(entry_id: str, data: dict[str, Any]) -> dict:
    init_db()
    existing = get_entry(entry_id)
    if not existing:
        raise KeyError("Not found")
    existing.update(normalize_entry_fields(data, partial=True))
    existing["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _lock:
        conn = _connect()
        try:
            conn.execute(
                """
                UPDATE entries SET
                  date=?, title=?, details=?, category=?, status=?,
                  tags=?, follow_up=?, owner=?, updated_at=?
                WHERE id=?
                """,
                (
                    existing["date"],
                    existing["title"],
                    existing["details"],
                    existing["category"],
                    existing["status"],
                    existing["tags"],
                    existing["follow_up"],
                    existing["owner"],
                    existing["updated_at"],
                    entry_id,
                ),
            )
            conn.commit()
        finally:
            conn.close()
    _after_write()
    return existing


def delete_entry(entry_id: str) -> None:
    init_db()
    with _lock:
        conn = _connect()
        try:
            cur = conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
            if cur.rowcount == 0:
                raise KeyError("Not found")
            conn.commit()
        finally:
            conn.close()
    _after_write()


def _after_write() -> None:
    """Refresh Excel mirror + rolling backup (best-effort)."""
    try:
        export_excel(EXCEL_PATH)
    except Exception as exc:
        # Don't fail the write if Excel is open/locked — DB is source of truth
        _record_last_export_error(str(exc))
        return
    try:
        create_backup(reason="auto")
    except Exception:
        pass


def _record_last_export_error(msg: str) -> None:
    try:
        init_db()
        with _lock:
            conn = _connect()
            try:
                _set_meta(conn, "last_export_error", msg[:500])
                conn.commit()
            finally:
                conn.close()
    except Exception:
        pass


def export_excel(path: Path | None = None) -> Path:
    """Write all entries to an .xlsx file. Raises StorageError on failure."""
    init_db()
    path = path or EXCEL_PATH
    entries = load_entries()
    # Newest rows at the top of the sheet (row 2 = most recent)
    ordered = sorted(entries, key=lambda e: (e["date"], e["created_at"]), reverse=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Log"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, e in enumerate(ordered, start=2):
        ws.cell(row=i, column=1, value=e["id"])
        ws.cell(row=i, column=2, value=e["date"])
        ws.cell(row=i, column=3, value=e["title"])
        ws.cell(row=i, column=4, value=e["details"])
        ws.cell(row=i, column=5, value=e["category"])
        ws.cell(row=i, column=6, value=e["status"])
        ws.cell(row=i, column=7, value=e.get("tags") or "")
        ws.cell(row=i, column=8, value=e.get("follow_up") or "")
        ws.cell(row=i, column=9, value=e.get("owner") or "")
        ws.cell(row=i, column=10, value=e["created_at"])
        ws.cell(row=i, column=11, value=e["updated_at"])
        ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True)
    widths = {
        "A": 36,
        "B": 12,
        "C": 40,
        "D": 70,
        "E": 14,
        "F": 12,
        "G": 18,
        "H": 12,
        "I": 14,
        "J": 20,
        "K": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".xlsx.tmp")
    try:
        wb.save(tmp)
        tmp.replace(path)
    except Exception as exc:
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass
        raise StorageError(
            f"Could not write Excel file (is it open in Excel?): {exc}"
        ) from exc
    return path


def backup_excel_copy(src: Path, label: str = "manual") -> Path | None:
    if not src.is_file():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = BACKUP_DIR / f"work_log-{stamp}-{label}.xlsx"
    shutil.copy2(src, dest)
    return dest


def create_backup(reason: str = "manual") -> dict[str, str]:
    """Snapshot SQLite + Excel into data/backups/. Returns paths."""
    init_db()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_reason = "".join(c if c.isalnum() or c in "-_" else "-" for c in reason)[:32]
    db_dest = BACKUP_DIR / f"worklog-{stamp}-{safe_reason}.db"
    xlsx_dest = BACKUP_DIR / f"work_log-{stamp}-{safe_reason}.xlsx"

    # Checkpoint WAL so backup is consistent
    with _lock:
        conn = _connect()
        try:
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            conn.commit()
        finally:
            conn.close()
        shutil.copy2(DB_PATH, db_dest)

    try:
        export_excel(xlsx_dest)
    except StorageError:
        # still keep DB backup
        if EXCEL_PATH.is_file():
            try:
                shutil.copy2(EXCEL_PATH, xlsx_dest)
            except OSError:
                xlsx_dest = Path("")

    _prune_backups()
    return {"db": str(db_dest), "xlsx": str(xlsx_dest) if xlsx_dest else ""}


def _prune_backups() -> None:
    if not BACKUP_DIR.is_dir():
        return
    for pattern in ("worklog-*.db", "work_log-*.xlsx"):
        files = sorted(
            BACKUP_DIR.glob(pattern),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for old in files[MAX_BACKUPS:]:
            try:
                old.unlink()
            except OSError:
                pass


def list_backups() -> list[dict[str, Any]]:
    init_db()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    items: list[dict[str, Any]] = []
    for p in sorted(BACKUP_DIR.glob("worklog-*.db"), key=lambda x: x.stat().st_mtime, reverse=True):
        items.append(
            {
                "name": p.name,
                "path": str(p),
                "size": p.stat().st_size,
                "mtime": datetime.fromtimestamp(p.stat().st_mtime).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                "kind": "db",
            }
        )
    return items[:MAX_BACKUPS]


def restore_backup(backup_name: str) -> None:
    """Restore SQLite from a backup file name under data/backups/."""
    init_db()
    # prevent path traversal
    name = Path(backup_name).name
    if not name.startswith("worklog-") or not name.endswith(".db"):
        raise ValueError("Invalid backup name")
    src = (BACKUP_DIR / name).resolve()
    if not str(src).startswith(str(BACKUP_DIR.resolve())):
        raise ValueError("Invalid backup path")
    if not src.is_file():
        raise FileNotFoundError("Backup not found")

    # Safety: backup current DB first
    create_backup(reason="pre-restore")

    with _lock:
        # close by replacing file
        dest = DB_PATH
        # remove wal/shm if present
        for suffix in ("-wal", "-shm"):
            side = Path(str(dest) + suffix)
            if side.exists():
                try:
                    side.unlink()
                except OSError:
                    pass
        shutil.copy2(src, dest)
        global _initialized
        _initialized = False
    init_db()
    try:
        export_excel(EXCEL_PATH)
    except StorageError:
        pass


def talking_points_text(start: str, end: str) -> str:
    entries = load_entries(start=start, end=end)
    # Newest first (same as UI lists)
    entries.sort(key=lambda e: (e["date"], e["created_at"]), reverse=True)
    lines = [f"Weekly summary {start} to {end}", ""]
    for e in entries:
        tags = f" [{e['tags']}]" if e.get("tags") else ""
        lines.append(f"- {e['date']}: {e['title']}{tags}")
        if e.get("details"):
            for dl in e["details"].splitlines():
                lines.append(f"    {dl}")
    if len(lines) == 2:
        lines.append("(no entries)")
    return "\n".join(lines) + "\n"


def stats() -> dict[str, Any]:
    init_db()
    total = count_entries()
    return {
        "total_entries": total,
        "db": str(DB_PATH),
        "excel": str(EXCEL_PATH),
        "excel_exists": EXCEL_PATH.is_file(),
        "backups": len(list_backups()),
    }
